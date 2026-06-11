# -*- coding: utf-8 -*-
"""Wizard import Excel/CSV → tạo Giấy báo có NH trích thu tự động.

Workflow:
  1. User upload file xlsx/csv
  2. Wizard parse → preview với match KW + kỳ + status
  3. User review, fix lỗi, click "Tạo giấy báo"
  4. Wizard tạo re.loan.bank.advice với lines (draft state) → mở form

Template format (xlsx hoặc csv):
  | Số HĐTD | Số khế ước | Số kỳ | Số tiền trích thu | Diễn giải |
  | HDTD001 | KU001      | 1     | 500000000         | Lãi T1    |
  | HDTD001 | KU002      |       | 200000000         |           |
"""
import base64
import csv
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None
    _logger.warning(
        "openpyxl not installed — xlsx import unavailable. "
        "pip install openpyxl")


COLUMNS = [
    ('credit_contract_code', 'Số HĐTD'),
    ('note_code',            'Số khế ước'),
    ('period_no',            'Số kỳ'),
    ('amount',               'Số tiền trích thu'),
    ('description',          'Diễn giải'),
]


class ReLoanBankAdviceImportWizard(models.TransientModel):
    _name = 're.loan.bank.advice.import.wizard'
    _description = 'Import Giấy báo có NH'

    # --- Upload step ---
    file_data = fields.Binary(
        string='File Excel/CSV', required=True,
        help='File xlsx hoặc csv theo template chuẩn (cột: Số HĐTD, '
             'Số khế ước, Số kỳ, Số tiền trích thu, Diễn giải).')
    file_name = fields.Char(string='Tên file')

    # --- Advice header ---
    partner_id = fields.Many2one(
        'res.partner', string='Ngân hàng', required=True,
        domain="[('is_bank', '=', True)]")
    bank_account_id = fields.Many2one(
        'res.partner.bank', string='TK trích thu',
        domain="[('partner_id', '=', company_partner_id),"
               " ('bank_id.partner_id', '=', partner_id)]",
        help='TK của công ty tại NH đã chọn — auto fill nếu chỉ có 1.')
    company_partner_id = fields.Many2one(
        'res.partner', readonly=True,
        compute='_compute_company_partner',
        help='Partner_id của company hiện tại — dùng để filter '
             'bank_account_id qua domain.')

    @api.depends_context('company')
    def _compute_company_partner(self):
        for rec in self:
            rec.company_partner_id = self.env.company.partner_id

    @api.onchange('partner_id')
    def _onchange_partner_autofill_bank_account(self):
        """Khi chọn NH, auto-pick TK của my company tại NH đó.

        - Nếu có đúng 1 TK → set sẵn
        - Nếu nhiều TK → để user chọn (dropdown filter sẽ rút gọn)
        - Nếu không có → clear field
        """
        self.bank_account_id = False
        if self.partner_id:
            accounts = self.env['res.partner.bank'].search([
                ('partner_id', '=', self.env.company.partner_id.id),
                ('bank_id.partner_id', '=', self.partner_id.id),
            ])
            if len(accounts) == 1:
                self.bank_account_id = accounts
    date_advice = fields.Date(
        string='Ngày NH trích thu', required=True,
        default=fields.Date.context_today)
    reference = fields.Char(string='Số chứng từ NH')
    description = fields.Text(string='Diễn giải')

    state = fields.Selection(
        [('upload',  'Upload file'),
         ('preview', 'Xem trước & sửa'),
         ('done',    'Đã tạo')],
        default='upload', required=True)

    preview_line_ids = fields.One2many(
        're.loan.bank.advice.import.line', 'wizard_id',
        string='Preview các dòng import')

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------
    def action_parse(self):
        """Parse file → fill preview_line_ids → state=preview."""
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Vui lòng upload file."))

        raw = base64.b64decode(self.file_data)
        rows = self._parse_file(raw, self.file_name or '')

        # Clear existing preview
        self.preview_line_ids.unlink()
        Line = self.env['re.loan.bank.advice.import.line']

        for idx, row in enumerate(rows, start=1):
            vals = self._build_preview_vals(idx, row)
            vals['wizard_id'] = self.id
            Line.create(vals)

        self.state = 'preview'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_create_advice(self):
        """Tạo re.loan.bank.advice + lines từ preview → mở form."""
        self.ensure_one()
        if self.state != 'preview':
            raise UserError(_("Phải parse file trước."))
        ok_lines = self.preview_line_ids.filtered(
            lambda l: l.status == 'ok')
        if not ok_lines:
            raise UserError(_(
                "Không có dòng nào hợp lệ để tạo giấy báo. "
                "Kiểm tra cột 'Trạng thái' để fix lỗi."))

        Advice = self.env['re.loan.bank.advice']
        advice = Advice.create({
            'partner_id': self.partner_id.id,
            'bank_account_id': (
                self.bank_account_id.id if self.bank_account_id else False),
            'date_advice': self.date_advice,
            'reference': self.reference or '',
            'description': self.description or '',
            'line_ids': [(0, 0, {
                'note_id': pl.note_id.id,
                'interest_line_id': (
                    pl.interest_line_id.id
                    if pl.interest_line_id else False),
                'amount': pl.amount,
                'description': pl.description or '',
            }) for pl in ok_lines],
        })
        self.state = 'done'
        return {
            'type': 'ir.actions.act_window',
            'name': _('Giấy báo có NH'),
            'res_model': 're.loan.bank.advice',
            'res_id': advice.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_download_template(self):
        """Trả về file xlsx template với sample."""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_(
                "Cần cài openpyxl trên server để generate template. "
                "Tạm thời copy template bằng tay theo cấu trúc:\n\n"
                "Số HĐTD | Số khế ước | Số kỳ | Số tiền trích thu | Diễn giải"))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Giấy báo NH'
        headers = [label for _, label in COLUMNS]
        ws.append(headers)
        # Sample rows
        ws.append(['HDTD_001', 'KU_001', 1, 500_000_000, 'Lãi T1/2026'])
        ws.append(['HDTD_001', 'KU_002', '', 200_000_000,
                   'Auto-allocate cũ→mới'])
        # Style header
        from openpyxl.styles import Font, PatternFill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(
                start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
        for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E'], 1):
            ws.column_dimensions[col_letter].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = base64.b64encode(buf.read())

        Attachment = self.env['ir.attachment']
        att = Attachment.create({
            'name': 'template_giay_bao_nh.xlsx',
            'type': 'binary',
            'datas': data,
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{att.id}?download=true',
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _parse_file(self, raw, file_name):
        """Detect xlsx/csv → return list of dict rows."""
        name = (file_name or '').lower()
        if name.endswith('.xlsx') or name.endswith('.xlsm'):
            return self._parse_xlsx(raw)
        if name.endswith('.csv') or name.endswith('.txt'):
            return self._parse_csv(raw)
        # Fallback: thử xlsx trước
        try:
            return self._parse_xlsx(raw)
        except Exception:
            return self._parse_csv(raw)

    def _parse_xlsx(self, raw):
        if not openpyxl:
            raise UserError(_(
                "openpyxl chưa cài. Upload file CSV thay vì xlsx, "
                "hoặc liên hệ admin cài: pip install openpyxl"))
        wb = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        return self._rows_to_dicts(rows)

    def _parse_csv(self, raw):
        text = raw.decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        return self._rows_to_dicts(rows)

    def _rows_to_dicts(self, rows):
        """Skip header row, map columns by position to COLUMNS keys."""
        result = []
        keys = [k for k, _ in COLUMNS]
        for row in rows[1:]:  # skip header
            if not any(c not in (None, '') for c in row):
                continue
            # Pad row to len(keys)
            padded = list(row) + [None] * (len(keys) - len(row))
            row_dict = {k: padded[i] for i, k in enumerate(keys)}
            result.append(row_dict)
        return result

    def _build_preview_vals(self, row_no, row):
        """Lookup KW + kỳ → return vals dict for preview line."""
        Note = self.env['re.loan.note']
        InterestLine = self.env['re.loan.note.interest.line']
        note_code = str(row.get('note_code') or '').strip()
        contract_code = str(row.get('credit_contract_code') or '').strip()
        period_raw = row.get('period_no')
        amount_raw = row.get('amount')
        description = str(row.get('description') or '').strip()

        vals = {
            'row_no': row_no,
            'note_code': note_code,
            'credit_contract_code': contract_code,
            'period_no_raw': str(period_raw) if period_raw not in (
                None, '') else False,
            'description': description,
        }

        # Parse amount
        try:
            amount = float(amount_raw) if amount_raw not in (None, '') else 0
        except (TypeError, ValueError):
            amount = 0
        vals['amount'] = amount

        # Validate amount
        if amount <= 0:
            vals['status'] = 'invalid_amount'
            vals['error_message'] = _("Số tiền không hợp lệ (= %s)") % amount
            return vals

        if not note_code:
            vals['status'] = 'no_match'
            vals['error_message'] = _("Thiếu Số khế ước")
            return vals

        # Match note by name
        notes = Note.search([('name', '=', note_code)], limit=2)
        if not notes:
            vals['status'] = 'no_match'
            vals['error_message'] = _(
                "Không tìm thấy KW '%s'") % note_code
            return vals
        if len(notes) > 1:
            vals['status'] = 'ambiguous'
            vals['error_message'] = _(
                "Số KW '%s' khớp %s records") % (note_code, len(notes))
            return vals
        note = notes[0]
        vals['note_id'] = note.id

        # Optional: validate credit_contract_code match
        if contract_code and note.credit_contract_id.name != contract_code:
            vals['status'] = 'contract_mismatch'
            vals['error_message'] = _(
                "Số HĐTD trong file '%(c)s' khác HĐTD của KW '%(a)s'",
                c=contract_code,
                a=note.credit_contract_id.name or '?')
            return vals

        # Optional: match kỳ
        if period_raw not in (None, ''):
            try:
                period_no = int(float(period_raw))
            except (TypeError, ValueError):
                vals['status'] = 'period_invalid'
                vals['error_message'] = _(
                    "Số kỳ không hợp lệ: '%s'") % period_raw
                return vals
            il = InterestLine.search([
                ('note_id', '=', note.id),
                ('period_no', '=', period_no),
            ], limit=1)
            if not il:
                vals['status'] = 'period_not_found'
                vals['error_message'] = _(
                    "KW '%(n)s' không có kỳ số %(p)s",
                    n=note_code, p=period_no)
                return vals
            vals['interest_line_id'] = il.id

        vals['status'] = 'ok'
        return vals


class ReLoanBankAdviceImportLine(models.TransientModel):
    _name = 're.loan.bank.advice.import.line'
    _description = 'Preview dòng import giấy báo'
    _order = 'row_no'

    wizard_id = fields.Many2one(
        're.loan.bank.advice.import.wizard',
        ondelete='cascade', required=True)
    row_no = fields.Integer(string='Dòng', readonly=True)

    # Raw values từ file
    credit_contract_code = fields.Char(string='Số HĐTD (file)')
    note_code = fields.Char(string='Số khế ước (file)')
    period_no_raw = fields.Char(string='Số kỳ (file)')
    amount = fields.Monetary(string='Số tiền')
    description = fields.Char(string='Diễn giải')

    # Matched records
    note_id = fields.Many2one(
        're.loan.note', string='Khế ước (match)',
        domain="[('name', '=', note_code)]")
    interest_line_id = fields.Many2one(
        're.loan.note.interest.line', string='Kỳ (match)',
        domain="[('note_id', '=', note_id)]")

    status = fields.Selection(
        [('ok',                'OK'),
         ('no_match',          'Không tìm thấy KW'),
         ('ambiguous',         'Trùng số KW'),
         ('contract_mismatch', 'HĐTD không khớp'),
         ('period_not_found',  'Kỳ không tồn tại'),
         ('period_invalid',    'Số kỳ sai định dạng'),
         ('invalid_amount',    'Số tiền không hợp lệ')],
        string='Trạng thái', default='ok')
    error_message = fields.Char(string='Lỗi', readonly=True)

    currency_id = fields.Many2one(
        'res.currency',
        default=lambda self: self.env.company.currency_id)
