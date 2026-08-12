# -*- coding: utf-8 -*-
"""Wizard nạp file Dự toán GXD — §5.4 + §3.5 của vn_cost_data_pattern.md.

Đây là thứ biến thư viện từ "gõ tay" thành "nạp file khách". Hai chữ ký kỹ
thuật, đều là bẫy quan sát được trong dữ liệu thật:

E1. **File GXD hỏng liên kết nội bộ.** GXD xuất `Target="NULL"` (chuỗi
    literal) trong `xl/drawings/_rels/*.rels` → openpyxl chết ngay:
    "There is no item named 'xl/drawings/NULL'". Excel lờ đi nên khách
    KHÔNG biết. Phải vá TRONG RAM trước khi đọc. XOÁ hẳn rel thì drawing
    mất rId1 → vẫn chết ⇒ phải TRỎ LẠI về ảnh có thật.

E2. **Đọc theo TÊN CỘT, không theo chỉ số.** GXD đổi layout giữa các phiên
    bản; một PDF/file chứa nhiều schema. Đọc theo index = đọc nhầm lặng lẽ.
    Layout đổi → báo lỗi RÕ, không đoán.

Bước này parse sheet giá tài nguyên (Nhan cong XD / Gia vat lieu HTXD /
Gia ca may XD) → tạo `rp.resource` (map `gxd_code`) + `rp.price.line`.
Parse "Don gia XD" (định mức) để sau.
"""
import base64
import io
import re
import zipfile

from odoo import _, fields, models
from odoo.exceptions import UserError


# ── E1: vá Target="NULL" trong RAM ──────────────────────────────────
def repair_gxd_xlsx(raw):
    """Nhận bytes .xlsx, trả bytes đã vá. Nếu không phải zip hợp lệ hoặc
    không có rel treo thì trả nguyên bytes gốc."""
    try:
        src = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        return raw, 0
    # tìm một ảnh media có thật để trỏ tới
    media = [n for n in src.namelist()
             if n.startswith('xl/media/') and not n.endswith('/')]
    target = ('../media/%s' % media[0].split('/')[-1]) if media else ''
    fixed = 0
    out = io.BytesIO()
    dst = zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED)
    for it in src.infolist():
        data = src.read(it.filename)
        if 'drawings/_rels' in it.filename and b'Target="NULL"' in data:
            if target:
                data = data.replace(b'Target="NULL"',
                                    ('Target="%s"' % target).encode())
            else:
                # không có ảnh nào để trỏ → gỡ cả entry rel treo
                data = re.sub(
                    rb'<Relationship[^>]*Target="NULL"[^>]*/>', b'', data)
            fixed += 1
        dst.writestr(it, data)
    dst.close()
    return out.getvalue(), fixed


# ── E2: map header theo TÊN, báo lỗi nếu thiếu cột bắt buộc ──────────
def header_map(ws, required):
    """Đọc dòng tiêu đề, trả dict {tên_chuẩn_hoá: chỉ_số_cột}. Tìm dòng
    header trong 8 dòng đầu (GXD có dòng tiêu đề công trình/hạng mục ở trên).
    """
    def norm(s):
        return re.sub(r'\s+', ' ', str(s or '').strip().lower())
    for r in range(1, 9):
        row = [norm(c.value) for c in ws[r]]
        hits = {}
        for key, aliases in required.items():
            for i, cell in enumerate(row):
                if any(a in cell for a in aliases):
                    hits[key] = i
                    break
        if len(hits) >= max(2, len(required) - 1):  # đủ phần lớn cột → là header
            return hits, r
    return {}, None


class RpGxdImport(models.TransientModel):
    _name = 'rp.gxd.import.wizard'
    _description = 'Nạp file Dự toán GXD'

    file = fields.Binary(string='File GXD (.xlsm / .xlsx)', required=True)
    filename = fields.Char(string='Tên file')
    publication_id = fields.Many2one(
        'rp.price.publication', string='Đưa giá vào công bố',
        help='Giá vật tư/NC/máy đọc được sẽ tạo dòng trong công bố này. '
             'Bỏ trống nếu chỉ muốn tạo tài nguyên (không giá).')
    result = fields.Text(string='Kết quả', readonly=True)

    def action_import(self):
        self.ensure_one()
        raw = base64.b64decode(self.file)
        raw, n_fixed = repair_gxd_xlsx(raw)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                io.BytesIO(raw), data_only=True, read_only=True)
        except Exception as e:  # noqa: BLE001
            raise UserError(_('Không mở được file: %s') % e)

        log = []
        if n_fixed:
            log.append(_('• Đã vá %s liên kết treo Target="NULL" '
                         '(lỗi phần mềm GXD xuất file — Excel lờ đi).')
                       % n_fixed)

        Res = self.env['rp.resource'].sudo()
        PL = self.env['rp.price.line'].sudo()
        n_res = n_price = n_skip = 0

        # ── các sheet giá tài nguyên GXD ──
        sheets = {
            'Nhan cong XD': ('labor', {
                'code': ['msvt'], 'name': ['loại nhân công', 'loai nhan cong'],
                'price': ['đơn giá nhân công', 'don gia nhan cong']}),
            'Gia vat lieu HTXD': ('material', {
                'code': ['msvt'], 'name': ['loại vật liệu', 'loai vat lieu'],
                'uom': ['đơn vị', 'don vi'],
                'price': ['giá vật liệu đến hiện trường',
                          'gia vat lieu den hien truong']}),
            'Gia ca may XD': ('machine', {
                'code': ['msvt'], 'name': ['loại máy', 'loai may'],
                'price': ['giá ca máy', 'gia ca may']}),
        }
        for sname, (rtype, cols) in sheets.items():
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            hmap, hrow = header_map(ws, cols)
            if 'code' not in hmap or 'name' not in hmap:
                log.append(_('• Sheet "%s": KHÔNG nhận ra cột (MSVT/tên). '
                             'Layout lạ — bỏ qua để không đọc nhầm.') % sname)
                continue
            for r in ws.iter_rows(min_row=hrow + 1, values_only=True):
                gxd = r[hmap['code']] if hmap['code'] < len(r) else None
                nm = r[hmap['name']] if hmap['name'] < len(r) else None
                if not gxd or not nm:
                    continue
                gxd = str(gxd).strip()
                nm = str(nm).strip()
                # mã tài nguyên GXD: 1-3 chữ + số (NC2357, M101.0106, V10135).
                # Loại junk header ("MÃ HIỆU", "[2]", "h") và mã định mức
                # (AB.11311 — có dấu chấm sau 2 chữ, không phải tài nguyên).
                if not re.match(r'^[A-Za-z]{1,3}\d', gxd):
                    continue
                res = Res.search([('gxd_code', '=', gxd)], limit=1)
                if not res:
                    res = Res.create({
                        'code': 'GXD.%s' % gxd, 'gxd_code': gxd, 'name': nm,
                        'resource_type': rtype,
                        'uom_id': self.env.ref('uom.product_uom_unit').id,
                    })
                    n_res += 1
                # giá
                if self.publication_id and 'price' in hmap:
                    pv = r[hmap['price']] if hmap['price'] < len(r) else None
                    if isinstance(pv, (int, float)) and pv > 0:
                        PL.create({
                            'publication_id': self.publication_id.id,
                            'resource_id': res.id, 'price': float(pv),
                            'vat_included': 'unknown',  # GXD không ghi rõ → không đoán
                            'price_basis': ('site_delivered'
                                            if rtype == 'material' else 'unknown'),
                        })
                        n_price += 1
                    else:
                        n_skip += 1

        log.append(_('• Tạo %s tài nguyên mới, %s dòng giá.') % (n_res, n_price))
        if n_skip:
            log.append(_('• %s dòng không có giá hợp lệ (bỏ qua).') % n_skip)
        if not (n_res or n_price):
            log.append(_('• Không đọc được sheet giá GXD nào '
                         '(Nhan cong XD / Gia vat lieu HTXD / Gia ca may XD).'))
        self.result = '\n'.join(log)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'rp.gxd.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
