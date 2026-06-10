import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ReFloorImportWizard(models.TransientModel):
    _name = 're.floor.import.wizard'
    _description = 'Real Estate Floor Import Wizard'

    project_id = fields.Many2one('re.project', string='Target Project', required=True)
    excel_file = fields.Binary(string='Excel File')
    file_name = fields.Char(string='File Name')
    update_existing = fields.Boolean(string='Update Existing Floors', default=True)

    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft')
    log_text = fields.Text(string='Log', readonly=True)
    success_count = fields.Integer(readonly=True)
    error_count = fields.Integer(readonly=True)
    skipped_count = fields.Integer(readonly=True)

    def action_download_template(self):
        if not openpyxl:
            raise UserError(_('openpyxl library not installed.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Floors'

        headers = ['building_code', 'floor_code', 'floor_number', 'floor_name', 'description']
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)

        samples = [
            ('B', 'B-B2', -2, 'Basement 2', ''),
            ('B', 'B-B1', -1, 'Basement 1', ''),
            ('B', 'B-GF', 0, 'Ground Floor', 'Lobby & retail'),
            ('B', 'B-01', 1, 'Floor 1', ''),
            ('B', 'B-02', 2, 'Floor 2', ''),
            ('B', 'B-PH', 99, 'Penthouse', 'Top floor luxury units'),
        ]
        for row_idx, sample in enumerate(samples, start=2):
            for col_idx, val in enumerate(sample, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'floor_import_template.xlsx',
            'datas': base64.b64encode(output.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('openpyxl library not installed.'))
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(_('Cannot read Excel file: %s') % str(e))

        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip().lower()] = col_idx

        required = {'building_code', 'floor_code', 'floor_number', 'floor_name'}
        missing = required - set(headers.keys())
        if missing:
            raise UserError(_('Missing required columns: %s') % ', '.join(sorted(missing)))

        success = skipped = 0
        errors = []
        log_lines = []

        buildings = {b.code: b for b in self.env['re.building'].search(
            [('project_id', '=', self.project_id.id)])}

        for row_idx in range(2, ws.max_row + 1):
            building_code_cell = ws.cell(row=row_idx, column=headers['building_code']).value
            floor_code_cell = ws.cell(row=row_idx, column=headers['floor_code']).value
            if not building_code_cell or not floor_code_cell:
                continue

            building_code = str(building_code_cell).strip()
            floor_code = str(floor_code_cell).strip()

            try:
                building = buildings.get(building_code)
                if not building:
                    raise UserError(_('Building "%s" not found') % building_code)

                floor_number_val = ws.cell(row=row_idx, column=headers['floor_number']).value
                try:
                    floor_number = int(floor_number_val)
                except (TypeError, ValueError):
                    raise UserError(_('floor_number must be integer'))

                floor_name = str(ws.cell(row=row_idx, column=headers['floor_name']).value or '').strip()
                if not floor_name:
                    raise UserError(_('floor_name cannot be empty'))

                description = ''
                if 'description' in headers:
                    desc_val = ws.cell(row=row_idx, column=headers['description']).value
                    description = str(desc_val) if desc_val else ''

                vals = {
                    'building_id': building.id, 'code': floor_code,
                    'name': floor_name, 'floor_number': floor_number,
                    'description': description,
                }

                existing = self.env['re.floor'].search([
                    ('building_id', '=', building.id),
                    ('code', '=', floor_code),
                ], limit=1)

                if existing:
                    if self.update_existing:
                        existing.write(vals)
                        log_lines.append(f'Row {row_idx}: UPDATE {building_code}/{floor_code}')
                        success += 1
                    else:
                        skipped += 1
                        log_lines.append(f'Row {row_idx}: SKIP {building_code}/{floor_code}')
                else:
                    self.env['re.floor'].create(vals)
                    log_lines.append(f'Row {row_idx}: CREATE {building_code}/{floor_code}')
                    success += 1
            except Exception as e:
                errors.append(f'Row {row_idx} ({building_code}/{floor_code}): {e}')

        self.write({
            'success_count': success,
            'error_count': len(errors),
            'skipped_count': skipped,
            'log_text': '\n'.join(log_lines + ['', '=== ERRORS ==='] + errors),
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 're.floor.import.wizard',
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }
