# -*- coding: utf-8 -*-
"""Sinh & đọc template Excel cho PoQ (4 bảng năng lực) và kế hoạch thi công.

Phong cách: nhà thầu tải template (đã bơm sẵn dữ liệu hiện có) → sửa trong
Excel → upload ngược để cập nhật. Dùng xlsxwriter (ghi) + openpyxl (đọc).
"""
import io

# Định nghĩa bảng: (sheet, model, [(header, field)]) — cột selection map nhãn↔key
POQ_SHEETS = [
    ('Nhân sự', 'cn.contractor.personnel', [
        ('Vị trí/chức danh', 'position'), ('Họ tên', 'name'),
        ('Trình độ/chuyên ngành', 'degree'), ('Số CCHN', 'cchn_number'),
        ('Hạng CCHN', 'cchn_grade'), ('Số năm KN', 'years_exp'),
        ('Số CT tương tự', 'similar_projects'), ('Huy động', 'mobilization'),
    ]),
    ('Thiết bị', 'cn.contractor.equipment', [
        ('Loại thiết bị', 'name'), ('Số lượng', 'quantity'),
        ('Công suất/thông số', 'capacity'), ('Hình thức', 'ownership'),
        ('Tình trạng', 'condition'),
    ]),
    ('Kinh nghiệm', 'cn.contractor.experience', [
        ('Tên hợp đồng', 'contract_name'), ('Chủ đầu tư', 'owner_name'),
        ('Giá trị', 'value'), ('Vai trò', 'role'),
        ('Loại công trình', 'work_type'), ('Cấp công trình', 'work_grade'),
        ('Quy mô/phần việc', 'scope'), ('Bắt đầu', 'date_start'),
        ('Kết thúc', 'date_end'), ('Đã nghiệm thu', 'accepted'),
    ]),
    ('Doanh thu', 'cn.contractor.revenue.year', [
        ('Năm', 'year'), ('Doanh thu (chưa VAT)', 'revenue'),
    ]),
]

SCHEDULE_HEADERS = [
    ('Mã WBS', 'wbs_code'), ('Tên công việc', 'name'),
    ('Bắt đầu', 'date_start'), ('Kết thúc', 'date_end'),
    ('% hoàn thành', 'progress_percent'), ('Là mốc', 'is_milestone'),
    ('Công việc trước', 'predecessors'),
]


def _sel_label(env, model, field, key):
    if not key:
        return ''
    sel = dict(env[model]._fields[field]._description_selection(env))
    return sel.get(key, key)


def _sel_key(env, model, field, label):
    """Nhãn (hoặc key) → key của selection; tolerant."""
    if not label:
        return False
    s = str(label).strip()
    sel = env[model]._fields[field]._description_selection(env)
    for k, lbl in sel:
        if s == k or s.lower() == lbl.lower():
            return k
    return False


def _cell(env, model, field, rec):
    """Giá trị 1 ô để ghi ra Excel (map selection → nhãn, date → str)."""
    f = env[model]._fields[field]
    val = rec[field]
    if f.type == 'selection':
        return _sel_label(env, model, field, val)
    if f.type == 'boolean':
        return 'x' if val else ''
    if f.type in ('date',):
        return val.strftime('%Y-%m-%d') if val else ''
    if f.type == 'many2one':
        return val.display_name if val else ''
    return val if val not in (False, None) else ''


def build_poq_template(partner):
    """Sinh workbook PoQ 4 sheet, bơm sẵn dữ liệu hiện có của nhà thầu."""
    import xlsxwriter
    env = partner.env
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True})
    hfmt = wb.add_format({'bold': True, 'bg_color': '#0e5563',
                          'font_color': 'white', 'border': 1})
    o2m = {
        'cn.contractor.personnel': partner.cn_personnel_ids,
        'cn.contractor.equipment': partner.cn_equipment_ids,
        'cn.contractor.experience': partner.cn_experience_ids,
        'cn.contractor.revenue.year': partner.cn_revenue_year_ids,
    }
    for sheet_name, model, cols in POQ_SHEETS:
        ws = wb.add_worksheet(sheet_name)
        for c, (header, _f) in enumerate(cols):
            ws.write(0, c, header, hfmt)
            ws.set_column(c, c, max(14, len(header) + 2))
        for r, rec in enumerate(o2m[model], start=1):
            for c, (_h, field) in enumerate(cols):
                ws.write(r, c, _cell(env, model, field, rec))
    wb.close()
    buf.seek(0)
    return buf.read()


def parse_poq(env, data):
    """Đọc workbook PoQ → dict {model: [vals,...]}. Bỏ qua sheet/hàng rỗng."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = {}
    by_sheet = {s: (model, cols) for s, model, cols in POQ_SHEETS}
    for ws in wb.worksheets:
        if ws.title not in by_sheet:
            continue
        model, cols = by_sheet[ws.title]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # khớp header thực tế với cột định nghĩa (theo thứ tự header)
        header = [str(c).strip().lower() if c is not None else ''
                  for c in rows[0]]
        idxmap = {}
        for field_i, (h, field) in enumerate(cols):
            for ci, hh in enumerate(header):
                if hh == h.lower():
                    idxmap[field] = ci
                    break
            else:
                idxmap[field] = field_i  # fallback theo vị trí
        recs = []
        for row in rows[1:]:
            vals = {}
            has_data = False
            for h, field in cols:
                ci = idxmap.get(field)
                raw = row[ci] if ci is not None and ci < len(row) else None
                if raw in (None, ''):
                    continue
                has_data = True
                vals[field] = _coerce(env, model, field, raw)
            if has_data:
                recs.append(vals)
        out[model] = recs
    return out


def _coerce(env, model, field, raw):
    f = env[model]._fields[field]
    if f.type == 'selection':
        return _sel_key(env, model, field, raw)
    if f.type == 'boolean':
        return str(raw).strip().lower() in ('1', 'x', 'true', 'có', 'co', 'yes')
    if f.type == 'integer':
        try:
            return int(float(str(raw).replace(',', '').strip()))
        except (ValueError, TypeError):
            return 0
    if f.type in ('float', 'monetary'):
        try:
            return float(str(raw).replace(',', '').strip())
        except (ValueError, TypeError):
            return 0.0
    if f.type == 'date':
        from .cn_schedule import _to_date
        return _to_date(raw) or False
    return str(raw).strip()


def build_schedule_template(bid):
    """Sinh workbook kế hoạch 1 sheet, bơm sẵn công việc hiện có."""
    import xlsxwriter
    env = bid.env
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True})
    hfmt = wb.add_format({'bold': True, 'bg_color': '#0e5563',
                          'font_color': 'white', 'border': 1})
    ws = wb.add_worksheet('Kế hoạch')
    for c, (header, _f) in enumerate(SCHEDULE_HEADERS):
        ws.write(0, c, header, hfmt)
        ws.set_column(c, c, max(14, len(header) + 2))
    for r, t in enumerate(bid.schedule_task_ids, start=1):
        for c, (_h, field) in enumerate(SCHEDULE_HEADERS):
            ws.write(r, c, _cell(env, 'cn.bid.schedule.task', field, t))
    wb.close()
    buf.seek(0)
    return buf.read()
