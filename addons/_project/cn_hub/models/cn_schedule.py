# -*- coding: utf-8 -*-
"""Kế hoạch/tiến độ thi công của nhà thầu trong hồ sơ dự thầu.

Nhà thầu nạp từ MS Project XML (MSPDI) hoặc template Excel → cn.bid.schedule.task,
sửa lại được sau import. Parse bê nguyên logic từ rp_schedule (không depend
chéo — cn_hub là stack Network riêng).
"""
import io
from datetime import datetime, date

from odoo import fields, models

MSPDI_NS = '{http://schemas.microsoft.com/project}'

# alias cột Excel (VN + EN) — khớp header không phân biệt hoa/thường
ALIASES = {
    'uid': {'uid', 'id', 'mã', 'ma', 'stt'},
    'wbs': {'wbs', 'mã wbs', 'ma wbs', 'wbs code'},
    'name': {'tên công việc', 'ten cong viec', 'công việc', 'cong viec',
             'task name', 'name', 'tên', 'ten', 'hạng mục', 'hang muc'},
    'start': {'bắt đầu', 'bat dau', 'ngày bắt đầu', 'start', 'start date'},
    'finish': {'kết thúc', 'ket thuc', 'ngày kết thúc', 'finish',
               'finish date', 'end'},
    'percent': {'% hoàn thành', '% hoan thanh', 'tiến độ', 'tien do',
                'percent', 'progress', '%'},
    'milestone': {'mốc', 'moc', 'milestone', 'là mốc', 'la moc'},
    'predecessors': {'công việc trước', 'cong viec truoc', 'trước',
                     'predecessors', 'predecessor'},
}


def _to_date(val):
    if not val:
        return False
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()[:10]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return False


def _to_percent(val):
    if val in (None, ''):
        return 0.0
    try:
        f = float(str(val).replace('%', '').strip())
    except (ValueError, TypeError):
        return 0.0
    return f * 100 if f <= 1 else f


def _truthy(val):
    return str(val).strip().lower() in ('1', 'true', 'x', 'yes', 'có', 'co')


def parse_excel(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    colmap = {}
    for key, aliases in ALIASES.items():
        for idx, h in enumerate(header):
            if h in aliases:
                colmap[key] = idx
                break
    if 'name' not in colmap:
        return []
    out = []
    for r in rows[1:]:
        name = r[colmap['name']] if colmap['name'] < len(r) else None
        if not name:
            continue

        def g(key):
            i = colmap.get(key)
            return r[i] if i is not None and i < len(r) else None
        out.append({
            'uid': str(g('uid')).strip() if g('uid') else False,
            'wbs': str(g('wbs')).strip() if g('wbs') else False,
            'name': str(name).strip(),
            'start': _to_date(g('start')),
            'finish': _to_date(g('finish')),
            'percent': _to_percent(g('percent')),
            'milestone': _truthy(g('milestone')),
            'predecessors': (str(g('predecessors')).strip()
                             if g('predecessors') else ''),
        })
    return out


def parse_mspdi(data):
    import xml.etree.ElementTree as ET
    root = ET.fromstring(data)
    out = []
    for t in root.iter(MSPDI_NS + 'Task'):
        def gt(tag):
            el = t.find(MSPDI_NS + tag)
            return el.text if el is not None else None
        name = gt('Name')
        if not name:
            continue
        preds = []
        for pl in t.iter(MSPDI_NS + 'PredecessorLink'):
            u = pl.find(MSPDI_NS + 'PredecessorUID')
            if u is not None and u.text:
                preds.append(u.text)
        out.append({
            'uid': (gt('UID') or '').strip() or False,
            'wbs': (gt('WBS') or gt('OutlineNumber') or '').strip() or False,
            'name': name.strip(),
            'start': _to_date(gt('Start')),
            'finish': _to_date(gt('Finish')),
            'percent': _to_percent(gt('PercentComplete')),
            'milestone': str(gt('Milestone')).strip() in ('1', 'true'),
            'predecessors': ','.join(preds),
        })
    return out


class CnBidScheduleTask(models.Model):
    _name = 'cn.bid.schedule.task'
    _description = 'Công việc kế hoạch thi công (hồ sơ dự thầu)'
    _order = 'bid_id, sequence, id'

    bid_id = fields.Many2one(
        'cn.bid', string='Hồ sơ dự thầu', required=True, ondelete='cascade',
        index=True)
    sequence = fields.Integer(default=10)
    external_uid = fields.Char(string='UID nguồn (MPP/Excel)', index=True)
    wbs_code = fields.Char(string='Mã WBS')
    name = fields.Char(string='Công việc', required=True)
    date_start = fields.Date(string='Bắt đầu')
    date_end = fields.Date(string='Kết thúc')
    progress_percent = fields.Float(string='% hoàn thành')
    is_milestone = fields.Boolean(string='Là mốc')
    predecessors = fields.Char(string='Công việc trước (UID)')
