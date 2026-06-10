import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


DIRECTION_MAPPING = {
    'n': 'n', 'north': 'n', 'bac': 'n', 'bắc': 'n',
    'ne': 'ne', 'northeast': 'ne', 'dong bac': 'ne', 'đông bắc': 'ne',
    'e': 'e', 'east': 'e', 'dong': 'e', 'đông': 'e',
    'se': 'se', 'southeast': 'se', 'dong nam': 'se', 'đông nam': 'se',
    's': 's', 'south': 's', 'nam': 's',
    'sw': 'sw', 'southwest': 'sw', 'tay nam': 'sw', 'tây nam': 'sw',
    'w': 'w', 'west': 'w', 'tay': 'w', 'tây': 'w',
    'nw': 'nw', 'northwest': 'nw', 'tay bac': 'nw', 'tây bắc': 'nw',
}

VIEW_MAPPING = {
    'lake': 'lake', 'river': 'river', 'sea': 'sea',
    'mountain': 'mountain', 'city': 'city', 'park': 'park',
    'pool': 'pool', 'internal': 'internal', 'garden': 'garden',
}


class ReUnitImportWizard(models.TransientModel):
    _name = 're.unit.import.wizard'
    _description = 'Real Estate Unit Import Wizard'

    project_id = fields.Many2one('re.project', string='Target Project', required=True)
    building_id = fields.Many2one(
        're.building', string='Target Building',
        domain="[('project_id', '=', project_id)]",
        help='Select to download a smart template pre-filled with one row per floor',
    )
    excel_file = fields.Binary(string='Excel File')
    file_name = fields.Char(string='File Name')

    state = fields.Selection([
        ('draft', 'Draft'),
        ('validated', 'Validated'),
        ('done', 'Done'),
    ], default='draft')
    log_text = fields.Text(string='Import Log', readonly=True)
    validation_log = fields.Text(string='Validation Log', readonly=True)
    success_count = fields.Integer(readonly=True)
    error_count = fields.Integer(readonly=True)
    skipped_count = fields.Integer(readonly=True)
    valid_row_count = fields.Integer(readonly=True,
        help='Number of rows that passed validation, ready to import.')

    @api.onchange('project_id')
    def _onchange_project_id(self):
        if self.building_id and self.building_id.project_id != self.project_id:
            self.building_id = False

    def action_download_template(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('openpyxl library not installed.'))
        if self.building_id:
            return self._download_smart_template()
        return self._download_generic_template()

    def _download_smart_template(self):
        building = self.building_id
        floors = building.floor_ids.sorted(key=lambda f: f.floor_number)
        if not floors:
            raise UserError(_(
                'Building "%s" has no floors yet.\n\nPlease generate or import floors first.'
            ) % building.display_name)

        wb = openpyxl.Workbook()

        # Instructions sheet (plain text only)
        ws_inst = wb.active
        ws_inst.title = 'Instructions'
        instructions = [
            'HUONG DAN DIEN TEMPLATE IMPORT UNITS',
            '',
            f'Project: {building.project_id.display_name}',
            f'Building: {building.display_name}',
            f'So tang da pre-fill: {len(floors)}',
            '',
            'CACH SU DUNG:',
            '1. Chuyen sang sheet "Units"',
            '2. Moi tang da duoc pre-fill 1 dong (building_code, floor_code, floor_name)',
            '3. Voi moi tang, COPY dong do thanh N dong theo so unit cua tang',
            '4. Dien unit_code va cac cot khac cho tung dong',
            '5. Save file Excel, upload va bam Import',
            '',
            'LUU Y:',
            '- unit_code phai UNIQUE trong ca du an',
            '- direction: East, North, Dong, Bac (deu duoc)',
            '- view_type: lake, river, sea, mountain, city, park, pool, internal, garden',
            '- area_net <= area_gross',
            '- Unit sau import se o trang thai Draft - chi available sau khi mo ban',
        ]
        for r_idx, text in enumerate(instructions, start=1):
            ws_inst.cell(row=r_idx, column=1, value=text)
        ws_inst.column_dimensions['A'].width = 90

        # Units sheet (plain, no styling)
        ws = wb.create_sheet('Units')
        headers = [
            'unit_code', 'building_code', 'floor_code', 'floor_name', 'unit_type_code',
            'area_net', 'area_gross', 'direction', 'view_type',
            'bedroom_count', 'bathroom_count', 'balcony_count',
            'original_price', 'notes',
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Pre-fill 1 row per floor
        row_idx = 2
        for floor in floors:
            ws.cell(row=row_idx, column=2, value=building.code)
            ws.cell(row=row_idx, column=3, value=floor.code)
            ws.cell(row=row_idx, column=4, value=floor.name)
            row_idx += 1

        widths = [16, 14, 12, 22, 14, 12, 12, 12, 12, 10, 10, 10, 16, 22]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        ws.freeze_panes = 'A2'

        return self._save_and_download(wb, f'unit_import_template_{building.code}.xlsx')

    def _download_generic_template(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Units'

        headers = [
            'unit_code', 'building_code', 'floor_code', 'unit_type_code',
            'area_net', 'area_gross', 'direction', 'view_type',
            'bedroom_count', 'bathroom_count', 'balcony_count',
            'original_price', 'notes',
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        sample = ['DSM-001', 'B', '02', 'STUDIO', 156.03, 162.50,
                  'East', 'lake', 1, 1, 1, 3914934843, '']
        for col_idx, val in enumerate(sample, start=1):
            ws.cell(row=2, column=col_idx, value=val)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

        return self._save_and_download(wb, 'unit_import_template.xlsx')

    def _save_and_download(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(output.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _validate_row(self, row_data, row_idx):
        """Validate a row WITHOUT creating anything.

        Returns dict {valid: bool, error: str|None, vals: dict|None}.
        Errors include suggested fixes (available codes).
        """
        unit_code = str(row_data.get('unit_code', '')).strip()
        if not unit_code:
            return {'valid': False, 'error': 'Empty unit_code', 'vals': None}

        building_code = str(row_data.get('building_code', '')).strip()
        floor_code = str(row_data.get('floor_code', '')).strip()
        unit_type_code = str(row_data.get('unit_type_code', '')).strip()

        # Building lookup
        building = self.env['re.building'].search([
            ('project_id', '=', self.project_id.id),
            ('code', '=', building_code),
        ], limit=1)
        if not building:
            available = self.env['re.building'].search([
                ('project_id', '=', self.project_id.id),
            ]).mapped('code')
            available_str = ', '.join(sorted(available)[:10]) or '(none)'
            return {
                'valid': False,
                'error': (
                    f'Building code "{building_code}" not found in project '
                    f'"{self.project_id.code}". Available: {available_str}'
                ),
                'vals': None,
            }

        # Floor lookup (within building)
        floor = self.env['re.floor'].search([
            ('building_id', '=', building.id),
            ('code', '=', floor_code),
        ], limit=1)
        if not floor:
            available = building.floor_ids.mapped('code')
            available_str = ', '.join(sorted(available)[:15]) or '(none)'
            return {
                'valid': False,
                'error': (
                    f'Floor code "{floor_code}" not found in building '
                    f'"{building.code}" (project "{self.project_id.code}"). '
                    f'Available floor codes in this building: {available_str}'
                ),
                'vals': None,
            }

        # Unit type lookup
        unit_type = self.env['re.unit.type'].search([
            ('code', '=', unit_type_code),
        ], limit=1)
        if not unit_type:
            available = self.env['re.unit.type'].search([]).mapped('code')
            available_str = ', '.join(sorted(available)[:15]) or '(none)'
            return {
                'valid': False,
                'error': (
                    f'Unit type code "{unit_type_code}" not found. '
                    f'Available: {available_str}'
                ),
                'vals': None,
            }

        # Duplicate unit_code check (within project)
        existing = self.env['re.unit'].search([
            ('project_id', '=', self.project_id.id),
            ('unit_code', '=', unit_code),
        ], limit=1)
        if existing:
            return {
                'valid': False,
                'error': (
                    f'Unit code "{unit_code}" already exists in project '
                    f'"{self.project_id.code}" (id={existing.id})'
                ),
                'vals': None,
            }

        # Numeric fields validation
        try:
            area_net = float(row_data.get('area_net') or 0)
            area_gross = float(row_data.get('area_gross') or 0)
            bedroom_count = int(row_data.get('bedroom_count') or 0)
            bathroom_count = int(row_data.get('bathroom_count') or 0)
            balcony_count = int(row_data.get('balcony_count') or 0)
            original_price = float(row_data.get('original_price') or 0)
        except (ValueError, TypeError) as e:
            return {
                'valid': False,
                'error': f'Numeric field parse error: {e}',
                'vals': None,
            }

        direction = self._map_direction(row_data.get('direction'))
        view_type = self._map_view(row_data.get('view_type'))

        vals = {
            'unit_code': unit_code,
            'project_id': self.project_id.id,
            'building_id': building.id,
            'floor_id': floor.id,
            'unit_type_id': unit_type.id,
            'area_net': area_net,
            'area_gross': area_gross,
            'bedroom_count': bedroom_count,
            'bathroom_count': bathroom_count,
            'balcony_count': balcony_count,
            'original_price': original_price,
            'direction': direction,
            'view_type': view_type,
            'notes': str(row_data.get('notes', '')).strip() or False,
            'state': 'draft',
        }

        return {'valid': True, 'error': None, 'vals': vals}

    def _import_one_row(self, row_data, log_lines, row_idx):
        """Legacy method kept for backward compat. Validates + creates in one call."""
        result = self._validate_row(row_data, row_idx)
        if not result['valid']:
            raise UserError(result['error'])
        unit = self.env['re.unit'].create(result['vals'])
        log_lines.append(f'Row {row_idx}: Created unit {unit.unit_code} (id={unit.id})')
        return unit

    def _read_excel(self):
        """Read Excel file and return (worksheet, headers_dict).
        headers_dict: {column_name_lower: col_idx}.
        """
        if not openpyxl:
            raise UserError(_('openpyxl library not installed.'))
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() in ('units', 'unit'):
                    ws = wb[sheet_name]
                    break
            if ws is None:
                ws = wb.active
        except Exception as e:
            raise UserError(_('Cannot read Excel file: %s') % str(e))

        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip().lower()] = col_idx

        required = {'unit_code', 'building_code', 'floor_code', 'unit_type_code'}
        missing = required - set(headers.keys())
        if missing:
            raise UserError(_(
                'Missing required columns: %s\n\n'
                'Required columns: %s'
            ) % (', '.join(sorted(missing)), ', '.join(sorted(required))))

        return ws, headers

    def action_set_to_draft(self):
        """Reset wizard to draft to re-upload or fix file."""
        self.ensure_one()
        self.write({
            'state': 'draft',
            'validation_log': False,
            'valid_row_count': 0,
            'error_count': 0,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 're.unit.import.wizard',
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    def action_validate(self):
        """Read file, validate every row, build report. Does NOT create units."""
        self.ensure_one()
        ws, headers = self._read_excel()

        valid_count = error_count = skipped = 0
        validation_lines = [
            '=== VALIDATION REPORT ===',
            f'Project: {self.project_id.code} - {self.project_id.name}',
            f'File: {self.file_name or "(uploaded)"}',
            '',
        ]
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            row_data = self._read_row(ws, row_idx, headers)
            if not row_data.get('unit_code'):
                skipped += 1
                continue

            result = self._validate_row(row_data, row_idx)
            if result['valid']:
                valid_count += 1
            else:
                error_count += 1
                errors.append(
                    f'Row {row_idx} (unit_code={row_data.get("unit_code")}): {result["error"]}'
                )

        # Build report
        validation_lines.append(f'✓ Valid rows: {valid_count}')
        validation_lines.append(f'✗ Invalid rows: {error_count}')
        validation_lines.append(f'○ Skipped (empty unit_code): {skipped}')
        validation_lines.append('')

        if errors:
            validation_lines.append('=== ERRORS ===')
            validation_lines.extend(errors)
            validation_lines.append('')
            validation_lines.append(
                'Fix these errors in the Excel file and re-upload, '
                'OR proceed to import only the valid rows.'
            )
        else:
            validation_lines.append('All rows passed validation. Click "Import" to proceed.')

        self.write({
            'valid_row_count': valid_count,
            'error_count': error_count,
            'skipped_count': skipped,
            'validation_log': '\n'.join(validation_lines),
            'state': 'validated',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 're.unit.import.wizard',
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    def action_import(self):
        """Import valid rows. Skip invalid ones (with error log)."""
        self.ensure_one()
        ws, headers = self._read_excel()

        success = skipped = 0
        errors = []
        log_lines = [
            '=== IMPORT LOG ===',
            f'Project: {self.project_id.code} - {self.project_id.name}',
            '',
        ]

        for row_idx in range(2, ws.max_row + 1):
            row_data = self._read_row(ws, row_idx, headers)
            if not row_data.get('unit_code'):
                skipped += 1
                continue

            result = self._validate_row(row_data, row_idx)
            if not result['valid']:
                errors.append(
                    f'Row {row_idx} (unit_code={row_data.get("unit_code")}): SKIPPED - {result["error"]}'
                )
                continue

            try:
                unit = self.env['re.unit'].create(result['vals'])
                success += 1
                log_lines.append(
                    f'Row {row_idx}: ✓ Created unit {unit.unit_code} (id={unit.id})'
                )
            except Exception as e:
                errors.append(
                    f'Row {row_idx} (unit_code={row_data.get("unit_code")}): CREATE FAILED - {e}'
                )

        log_lines.append('')
        log_lines.append(f'=== SUMMARY ===')
        log_lines.append(f'✓ Imported: {success}')
        log_lines.append(f'✗ Errors: {len(errors)}')
        log_lines.append(f'○ Skipped (empty): {skipped}')

        if errors:
            log_lines.append('')
            log_lines.append('=== ERRORS ===')
            log_lines.extend(errors)

        self.write({
            'success_count': success,
            'error_count': len(errors),
            'skipped_count': skipped,
            'log_text': '\n'.join(log_lines),
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 're.unit.import.wizard',
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    def _read_row(self, ws, row_idx, headers):
        row_data = {}
        for col_name, col_idx in headers.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data[col_name] = cell_value
        return row_data

    def _map_direction(self, value):
        if not value:
            return False
        return DIRECTION_MAPPING.get(str(value).strip().lower(), False)

    def _map_view(self, value):
        if not value:
            return False
        return VIEW_MAPPING.get(str(value).strip().lower(), False)
