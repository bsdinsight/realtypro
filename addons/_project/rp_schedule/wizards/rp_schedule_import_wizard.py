# -*- coding: utf-8 -*-
"""Import lịch thi công từ Excel hoặc MS Project XML (MSPDI) → project.task.

- Excel: header linh hoạt (alias VN/EN), cột WBS/Tên/Bắt đầu/Kết thúc/%/
  Milestone/UID/Predecessors.
- MS Project XML (MSPDI, .xml Save As từ MS Project): parse <Task>.
Re-import idempotent theo external_uid trong phạm vi HĐ (cập nhật %/ngày).
"""
import base64
import io
from datetime import datetime, date

from odoo import _, api, fields, models
from odoo.exceptions import UserError

MSPDI_NS = '{http://schemas.microsoft.com/project}'

# Alias header Excel (chữ thường, không dấu tùy biến người dùng nhập)
ALIASES = {
    'wbs': ['wbs', 'ma wbs', 'mã wbs', 'outline', 'outline number', 'stt'],
    'name': ['name', 'task name', 'ten', 'tên', 'ten cong viec',
             'tên công việc', 'cong viec', 'công việc', 'task'],
    'start': ['start', 'bat dau', 'bắt đầu', 'ngay bat dau', 'ngày bắt đầu',
              'planned start', 'start date'],
    'finish': ['finish', 'end', 'ket thuc', 'kết thúc', 'ngay ket thuc',
               'ngày kết thúc', 'finish date'],
    'percent': ['% complete', 'percent complete', '%', '% hoan thanh',
                '% hoàn thành', 'hoan thanh', 'hoàn thành', 'percent',
                'tien do', 'tiến độ'],
    'milestone': ['milestone', 'moc', 'mốc', 'la moc', 'là mốc'],
    'uid': ['uid', 'id', 'unique id', 'ma', 'mã'],
    'predecessors': ['predecessors', 'predecessor', 'cong viec truoc',
                     'công việc trước', 'truoc', 'trước'],
}


class RpScheduleImportWizard(models.TransientModel):
    _name = 'rp.schedule.import.wizard'
    _description = 'Import lịch thi công (Excel / MS Project XML)'

    contract_id = fields.Many2one(
        'rp.contract', string='HĐ nhà thầu', required=True)
    file = fields.Binary(string='File (.xlsx / .xml)', required=True)
    filename = fields.Char(string='Tên file')
    default_structure_id = fields.Many2one(
        'rp.structure', string='Hạng mục mặc định',
        help='Gán cho các công việc chưa map hạng mục (tùy chọn).')

    # ---------- helpers ----------
    @staticmethod
    def _to_date(val):
        if not val:
            return False
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        s = str(val).strip()
        if not s:
            return False
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d', '%d/%m/%Y',
                    '%m/%d/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(s[:19] if 'T' in s else s,
                                         fmt).date()
            except ValueError:
                continue
        return False

    @staticmethod
    def _to_percent(val):
        if val in (None, '', False):
            return 0.0
        try:
            f = float(val)
        except (ValueError, TypeError):
            return 0.0
        if 0.0 < f <= 1.0:      # Excel percent-format (0.5 = 50%)
            f *= 100.0
        return max(0.0, min(100.0, f))

    @staticmethod
    def _truthy(val):
        return str(val).strip().lower() in (
            '1', 'true', 'yes', 'x', 'có', 'co', 'milestone')

    # ---------- parse ----------
    def _parse_excel(self, data):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # tìm hàng header
        header = [str(c).strip().lower() if c is not None else ''
                  for c in rows[0]]
        colmap = {}
        for key, aliases in ALIASES.items():
            for idx, h in enumerate(header):
                if h in aliases:
                    colmap[key] = idx
                    break
        if 'name' not in colmap:
            raise UserError(_(
                'Không tìm thấy cột "Tên công việc" trong file Excel. '
                'Header cần có 1 trong: %s', ', '.join(ALIASES['name'])))
        out = []
        for r in rows[1:]:
            name = r[colmap['name']] if colmap['name'] < len(r) else None
            if not name:
                continue

            def g(key):
                i = colmap.get(key)
                return r[i] if i is not None and i < len(r) else None
            out.append({
                'uid': str(g('uid')).strip() if g('uid') else False,
                'wbs': str(g('wbs')).strip() if g('wbs') else False,
                'name': str(name).strip(),
                'start': self._to_date(g('start')),
                'finish': self._to_date(g('finish')),
                'percent': self._to_percent(g('percent')),
                'milestone': self._truthy(g('milestone')),
                'predecessors': (str(g('predecessors')).strip()
                                 if g('predecessors') else ''),
            })
        return out

    def _parse_mspdi(self, data):
        import xml.etree.ElementTree as ET
        root = ET.fromstring(data)
        out = []
        for t in root.iter(MSPDI_NS + 'Task'):
            def gt(tag):
                el = t.find(MSPDI_NS + tag)
                return el.text if el is not None else None
            name = gt('Name')
            if not name:
                continue
            preds = []
            for pl in t.iter(MSPDI_NS + 'PredecessorLink'):
                u = pl.find(MSPDI_NS + 'PredecessorUID')
                if u is not None and u.text:
                    preds.append(u.text)
            out.append({
                'uid': (gt('UID') or '').strip() or False,
                'wbs': (gt('WBS') or gt('OutlineNumber') or '').strip()
                or False,
                'name': name.strip(),
                'start': self._to_date(gt('Start')),
                'finish': self._to_date(gt('Finish')),
                'percent': self._to_percent(gt('PercentComplete')),
                'milestone': str(gt('Milestone')).strip() in ('1', 'true'),
                'predecessors': ','.join(preds),
            })
        return out

    # ---------- import ----------
    def action_import(self):
        self.ensure_one()
        c = self.contract_id
        proj = c._get_or_create_schedule_project()
        data = base64.b64decode(self.file)
        name = (self.filename or '').lower()
        if name.endswith('.xml'):
            rows = self._parse_mspdi(data)
        elif name.endswith(('.xlsx', '.xlsm')):
            rows = self._parse_excel(data)
        else:
            # thử Excel trước, fallback XML
            try:
                rows = self._parse_excel(data)
            except Exception:
                rows = self._parse_mspdi(data)
        if not rows:
            raise UserError(_('File không có dòng công việc nào đọc được.'))

        Task = self.env['project.task']
        existing = {t.external_uid: t for t in c.task_ids if t.external_uid}
        uid_to_task = {}
        created = updated = 0
        for row in rows:
            vals = {
                'name': row['name'],
                'rp_contract_id': c.id,
                'project_id': proj.id,
                'wbs_code': row['wbs'],
                'planned_start': row['start'],
                'planned_end': row['finish'],
                'progress_percent': row['percent'],
                'is_milestone': row['milestone'],
                'external_uid': row['uid'],
            }
            if self.default_structure_id:
                vals['rp_structure_id'] = self.default_structure_id.id
            task = existing.get(row['uid']) if row['uid'] else False
            if task:
                task.write(vals)
                updated += 1
            else:
                task = Task.create(vals)
                created += 1
            if row['uid']:
                uid_to_task[row['uid']] = task

        # pass 2: link predecessors theo UID
        for row in rows:
            if not row['predecessors'] or not row['uid']:
                continue
            task = uid_to_task.get(row['uid'])
            if not task:
                continue
            preds = [uid_to_task[u.strip()]
                     for u in row['predecessors'].replace(';', ',').split(',')
                     if u.strip() in uid_to_task]
            if preds:
                task.predecessor_ids = [(6, 0, [p.id for p in preds])]

        c.message_post(body=_(
            'Import lịch thi công: %(c)s công việc mới, %(u)s cập nhật '
            '(từ %(f)s).', c=created, u=updated, f=self.filename or 'file'))
        return {
            'type': 'ir.actions.act_window',
            'name': _('Lịch thi công — %s', c.name),
            'res_model': 'project.task',
            'view_mode': 'list,form',
            'domain': [('rp_contract_id', '=', c.id)],
            'context': {'default_rp_contract_id': c.id,
                        'default_project_id': proj.id},
        }
